"""Bygger prices.json: arlige priser for et forenklet inntektsestimat.

Kilder (alle apne, ingen API-nokkel):
  - Verdensbanken "Pink Sheet" (CMO-Historical-Data-Monthly.xlsx)
      Crude oil, Brent  ($/bbl)      -> kolonne C, fra 1960
      Natural gas, Europe ($/mmbtu)  -> kolonne I, fra 1960
  - EIA Europe Brent Spot (RBRTE, manedlig) for ferske maneder Verdensbanken mangler
  - Norges Bank EXR A.USD.NOK.SP    -> arlig gjennomsnittskurs

MERK: dette gir et GROVT ANSLAG pa bruttoverdien av produsert petroleum.
Det er ikke faktisk salgsinntekt. Reelle kontrakter avviker fra spot, gass
selges pa lange kontrakter, og NGL/kondensat prises med egne differanser.
"""
import datetime
import json
import os
import re
import urllib.request

WB_XLSX = "_wb.xlsx"
WB_URL = ("https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025"
          "/related/CMO-Historical-Data-Monthly.xlsx")
NB_URL = ("https://data.norges-bank.no/api/data/EXR/A.USD.NOK.SP"
          "?format=csv&startPeriod=1971")
NB_URL_M = ("https://data.norges-bank.no/api/data/EXR/M.USD.NOK.SP"
            "?format=csv&startPeriod=2026-01")
EIA_URL = "https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx?n=PET&s=RBRTE&f=M"

UA = {"User-Agent": "Mozilla/5.0 (sokkel-mvp; apne data)"}


def get(url: str, timeout: int = 120) -> str:
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", errors="replace")


def world_bank():
    """Arlige snitt for Brent og europeisk gass."""
    import openpyxl
    if not os.path.exists(WB_XLSX):
        print("laster Pink Sheet ...")
        urllib.request.urlretrieve(WB_URL, WB_XLSX)
    wb = openpyxl.load_workbook(WB_XLSX, data_only=True)
    ws = wb["Monthly Prices"]

    oil, gas = {}, {}
    for r in range(7, ws.max_row + 1):
        tag = ws.cell(r, 1).value
        if not tag or not re.match(r"^\d{4}M\d{2}$", str(tag)):
            continue
        yr = int(str(tag)[:4])
        for col, store in ((3, oil), (9, gas)):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                store.setdefault(yr, []).append(float(v))

    avg = lambda d: {y: round(sum(v) / len(v), 3) for y, v in d.items()}
    o, g = avg(oil), avg(gas)
    print(f"  Verdensbanken: olje {min(o)}-{max(o)}, gass {min(g)}-{max(g)}")
    return o, g


def eia_recent():
    """Manedlige Brent-priser fra EIA, for ar Verdensbanken ikke dekker enda."""
    try:
        html = get(EIA_URL)
    except Exception as e:
        print(f"  EIA utilgjengelig ({e}) - hopper over")
        return {}
    out = {}
    # Radene er delt pa \r, sa linjebasert parsing feiler. Ta <tr>-blokker i stedet.
    for row in re.findall(r"<tr\b.*?</tr>", html, re.S | re.I):
        cells = [re.sub(r"<[^>]+>", "", c).replace("&nbsp;", " ").strip()
                 for c in re.findall(r"<td\b.*?</td>", row, re.S | re.I)]
        if not cells or not re.match(r"^(19|20)\d{2}$", cells[0]):
            continue
        vals = []
        for c in cells[1:13]:
            try:
                vals.append(float(c))
            except ValueError:
                pass
        if vals:
            out[int(cells[0])] = (round(sum(vals) / len(vals), 3), len(vals))
    if out:
        y = max(out)
        print(f"  EIA Brent: {min(out)}-{y}  (siste: {out[y][0]} $/fat, {out[y][1]} mnd)")
    else:
        print("  EIA: fant ingen rader")
    return out


def norges_bank():
    """Arlig snittkurs. Inneverende ar beregnes fra manedstall."""
    out = {}
    try:
        txt = get(NB_URL)
    except Exception as e:
        print(f"  Norges Bank utilgjengelig ({e})")
        return out
    for line in txt.splitlines()[1:]:
        p = line.split(";")
        if len(p) < 2:
            continue
        yr, val = p[-2].strip(), p[-1].strip().replace(",", ".")
        if re.match(r"^\d{4}$", yr):
            try:
                out[int(yr)] = round(float(val), 4)
            except ValueError:
                pass

    # inneverende ar mangler i den arlige serien - regn snitt av manedene
    try:
        m = get(NB_URL_M)
        buckets = {}
        for line in m.splitlines()[1:]:
            p = line.split(";")
            if len(p) < 2:
                continue
            per, val = p[-2].strip(), p[-1].strip().replace(",", ".")
            if re.match(r"^\d{4}-\d{2}$", per):
                try:
                    buckets.setdefault(int(per[:4]), []).append(float(val))
                except ValueError:
                    pass
        for y, vals in buckets.items():
            if y not in out and vals:
                out[y] = round(sum(vals) / len(vals), 4)
                print(f"  + USD/NOK {y} fra manedstall: {out[y]} ({len(vals)} mnd)")
    except Exception as e:
        print(f"  Norges Bank manedstall utilgjengelig ({e})")

    if out:
        print(f"  Norges Bank USD/NOK: {min(out)}-{max(out)}  (siste: {out[max(out)]})")
    return out


def eia_daily():
    """Siste daglige Brent-notering. Brukes til sanntidstelleren.

    Radetiketten er en ukesperiode, f.eks. "2026 Aug-24 to Aug-28".
    Verdiene er paafolgende virkedager, saa vi regner oss fram til den
    faktiske datoen for siste noterte pris.
    """
    MND = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
           "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
    NO = {1: "januar", 2: "februar", 3: "mars", 4: "april", 5: "mai", 6: "juni",
          7: "juli", 8: "august", 9: "september", 10: "oktober",
          11: "november", 12: "desember"}
    try:
        html = get("https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx?n=PET&s=RBRTE&f=D")
    except Exception as e:
        print(f"  EIA daglig utilgjengelig ({e})")
        return None

    siste = None
    for row in re.findall(r"<tr\b.*?</tr>", html, re.S | re.I):
        cells = [re.sub(r"<[^>]+>", "", c).replace("&nbsp;", " ").strip()
                 for c in re.findall(r"<td\b.*?</td>", row, re.S | re.I)]
        m = re.match(r"^(\d{4})\s+(\w{3})-(\d{1,2})", cells[0] if cells else "")
        if not m:
            continue
        verdier = []
        for c in cells[1:]:
            try:
                verdier.append(float(c))
            except ValueError:
                pass
        if not verdier:
            continue
        aar, mnd, dag = int(m.group(1)), MND.get(m.group(2)), int(m.group(3))
        if not mnd:
            continue
        try:
            d = datetime.date(aar, mnd, dag) + datetime.timedelta(days=len(verdier) - 1)
            tekst = f"{d.day}. {NO[d.month]} {d.year}"
        except ValueError:
            tekst = cells[0]
        siste = (verdier[-1], tekst)

    if siste:
        print(f"  EIA Brent daglig: {siste[0]} $/fat ({siste[1]})")
    return siste


def nb_daily():
    """Siste daglige USD/NOK fra Norges Bank."""
    try:
        txt = get("https://data.norges-bank.no/api/data/EXR/B.USD.NOK.SP"
                  "?format=csv&lastNObservations=1")
    except Exception as e:
        print(f"  Norges Bank daglig utilgjengelig ({e})")
        return None
    for line in txt.splitlines()[1:]:
        p = line.split(";")
        if len(p) < 2:
            continue
        try:
            v = float(p[-1].strip().replace(",", "."))
            print(f"  Norges Bank USD/NOK daglig: {v} ({p[-2].strip()})")
            return (v, p[-2].strip())
        except ValueError:
            pass
    return None


# 1 MMBtu = 1.055056 GJ, 1 MWh = 3.6 GJ  =>  1 MWh = 3.6/1.055056 MMBtu
MMBTU_PER_MWH = 3.6 / 1.055056

NO_MND = {1: "januar", 2: "februar", 3: "mars", 4: "april", 5: "mai", 6: "juni",
          7: "juli", 8: "august", 9: "september", 10: "oktober",
          11: "november", 12: "desember"}


def ttf_daily():
    """Siste TTF-notering (europeisk referansepris for gass, Rotterdam-borsen
    ICE Endex), regnet om fra EUR/MWh til USD/MMBtu.

    FRED var utilgjengelig herfra (se todo), og Verdensbankens Pink Sheet
    henger over ett ar etter for gass. Yahoo Finance sin (uoffisielle, men
    apne og nokkelfrie) chart-API for TTF-terminkontrakten "TTF=F" gir en
    fersk dagsnotering uten registrering. Kun kjort fra byggeskriptet
    (server-side/Python) - dette er IKKE tilgjengelig fra nettleseren, siden
    Yahoo ikke setter CORS-headere for vilkarlige opphav.
    """
    try:
        j = json.loads(get("https://query1.finance.yahoo.com/v8/finance/chart/TTF=F"
                            "?range=5d&interval=1d"))
        meta = j["chart"]["result"][0]["meta"]
        eur_per_mwh = meta["regularMarketPrice"]
        ts = meta.get("regularMarketTime")
    except Exception as e:
        print(f"  TTF (Yahoo) utilgjengelig ({e})")
        return None

    eurusd = 1.08  # forsiktig anslag hvis kursoppslaget under feiler
    try:
        j2 = json.loads(get("https://query1.finance.yahoo.com/v8/finance/chart/EURUSD=X"
                             "?range=5d&interval=1d"))
        eurusd = j2["chart"]["result"][0]["meta"]["regularMarketPrice"]
    except Exception as e:
        print(f"  EUR/USD (Yahoo) utilgjengelig ({e}) - bruker {eurusd} som anslag")

    usd_per_mmbtu = round(eur_per_mwh * eurusd / MMBTU_PER_MWH, 3)
    if ts:
        d = datetime.date.fromtimestamp(ts)
        tekst = f"{d.day}. {NO_MND[d.month]} {d.year}"
    else:
        tekst = "ukjent dato"
    print(f"  TTF dagsnotering: {eur_per_mwh} EUR/MWh × {eurusd} USD/EUR "
          f"= {usd_per_mmbtu} $/MMBtu ({tekst})")
    return (usd_per_mmbtu, tekst)


def main():
    print("bygger prisdata ...")
    oil, gas = world_bank()
    eia = eia_recent()
    fx = norges_bank()
    dag_olje = eia_daily()
    dag_fx = nb_daily()
    dag_gass = ttf_daily()

    # EIA fyller inn ar Verdensbanken mangler (typisk inneverende ar)
    partial = {}
    for y, (v, nmonths) in eia.items():
        if y not in oil:
            oil[y] = v
            if nmonths < 12:
                partial[y] = nmonths
            print(f"  + Brent {y} fra EIA: {v} $/fat ({nmonths} mnd)")

    years = sorted(set(oil) | set(gas) | set(fx))

    def carry(d, yrs):
        """Fyll hull framover med siste kjente verdi."""
        out, last = {}, None
        for y in yrs:
            if y in d:
                last = d[y]
            if last is not None:
                out[y] = last
        return out

    oil_f, gas_f, fx_f = carry(oil, years), carry(gas, years), carry(fx, years)

    data = {
        "note": "Grovt anslag. Spotpriser, ikke faktiske salgsinntekter.",
        "sources": {
            "oil": "Verdensbanken Pink Sheet (Crude oil, Brent) + EIA RBRTE",
            "gas": "Verdensbanken Pink Sheet (Natural gas, Europe)"
                   + (" + TTF-terminkontrakt (dagsnotering)" if dag_gass else ""),
            "fx": "Norges Bank EXR A.USD.NOK.SP",
        },
        "units": {"oil": "USD/fat", "gas": "USD/MMBtu", "fx": "NOK per USD"},
        "lastFullYear": max(y for y in years if y not in partial),
        "partialYears": partial,
        "carriedGasFrom": max(gas) if gas else None,
        # ferskeste noteringer - brukes til sanntidstelleren
        "spot": {
            "brent": dag_olje[0] if dag_olje else None,
            "brentDato": dag_olje[1] if dag_olje else None,
            "fx": dag_fx[0] if dag_fx else None,
            "fxDato": dag_fx[1] if dag_fx else None,
            "gass": dag_gass[0] if dag_gass else (gas_f.get(max(gas)) if gas else None),
            "gassDato": dag_gass[1] if dag_gass else (f"arssnitt {max(gas)}" if gas else None),
        },
        "years": years,
        "oil": [oil_f.get(y) for y in years],
        "gas": [gas_f.get(y) for y in years],
        "fx": [fx_f.get(y) for y in years],
    }
    with open("prices.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    print(f"\nprices.json: {round(os.path.getsize('prices.json') / 1024, 1)} KB, "
          f"{len(years)} ar ({years[0]}-{years[-1]})")
    for y in (1980, 2000, 2014, 2022, 2024, 2025, 2026):
        if y in oil_f:
            print(f"  {y}: Brent {oil_f.get(y)} $/fat | gass {gas_f.get(y)} $/MMBtu "
                  f"| USD/NOK {fx_f.get(y)}")


if __name__ == "__main__":
    main()
